"""Controlador de reportes deportivos."""

import logging
from datetime import date, datetime
from io import BytesIO
from typing import List, Optional
import uuid

from sqlalchemy.orm import Session

from app.dao.report_dao import ReportDAO
from app.schemas.report_schema import (
    ReportFilter,
    ReportGenerationResponse,
    ReportSummary,
    AttendanceReportData,
    EvaluationReportData,
    TestResultReportData,
)
from app.utils.exceptions import ValidationException

logger = logging.getLogger(__name__)


class ReportController:
    """Controlador de reportes deportivos."""

    def __init__(self):
        self.report_dao = ReportDAO()

    def validate_filters(self, filters: ReportFilter) -> None:
        """
        Valida los filtros de reporte.

        Args:
            filters: Filtros a validar

        Raises:
            ValidationException: Si los filtros son inválidos
        """
        # Validar rango de fechas
        if filters.start_date and filters.end_date:
            if filters.start_date > filters.end_date:
                raise ValidationException(
                    "start_date debe ser menor o igual a end_date"
                )

        # Validar que al menos un filtro de inclusión esté activo
        if not any(
            [
                filters.include_attendance,
                filters.include_evaluations,
                filters.include_tests,
            ]
        ):
            raise ValidationException(
                "Al menos uno de los filtros de inclusión debe estar activo"
            )

    def collect_report_data(
        self,
        db: Session,
        filters: ReportFilter,
    ) -> dict:
        """
        Recolecta todos los datos necesarios para el reporte.

        Args:
            db: Sesión de BD
            filters: Filtros del reporte

        Returns:
            Diccionario con datos del reporte
        """
        self.validate_filters(filters)

        # Obtener atletas
        athletes = self.report_dao.get_athletes_for_report(
            db=db,
            club_id=filters.club_id,
            athlete_id=filters.athlete_id,
        )

        if not athletes:
            raise ValidationException("No se encontraron atletas con los filtros especificados")

        athlete_ids = [athlete.id for athlete in athletes]

        report_data = {
            "athletes": athletes,
            "attendance": [],
            "evaluations": [],
            "tests": {
                "sprint": [],
                "endurance": [],
                "yoyo": [],
                "technical": [],
            },
            "statistics": {},
        }

        # Recolectar datos de asistencia
        if filters.include_attendance:
            report_data["attendance"] = self.report_dao.get_attendance_records(
                db=db,
                athlete_ids=athlete_ids,
                start_date=filters.start_date,
                end_date=filters.end_date,
            )

        # Recolectar datos de evaluaciones
        if filters.include_evaluations:
            report_data["evaluations"] = self.report_dao.get_evaluations(
                db=db,
                athlete_ids=athlete_ids,
                start_date=filters.start_date,
                end_date=filters.end_date,
            )

        # Recolectar resultados de pruebas
        if filters.include_tests:
            report_data["tests"]["sprint"] = self.report_dao.get_sprint_tests(
                db=db,
                athlete_ids=athlete_ids,
                start_date=filters.start_date,
                end_date=filters.end_date,
            )

            report_data["tests"]["endurance"] = self.report_dao.get_endurance_tests(
                db=db,
                athlete_ids=athlete_ids,
                start_date=filters.start_date,
                end_date=filters.end_date,
            )

            report_data["tests"]["yoyo"] = self.report_dao.get_yoyo_tests(
                db=db,
                athlete_ids=athlete_ids,
                start_date=filters.start_date,
                end_date=filters.end_date,
            )

            report_data["tests"]["technical"] = (
                self.report_dao.get_technical_assessments(
                    db=db,
                    athlete_ids=athlete_ids,
                    start_date=filters.start_date,
                    end_date=filters.end_date,
                )
            )

        # Obtener estadísticas
        report_data["statistics"] = self.report_dao.get_report_statistics(
            db=db,
            athlete_ids=athlete_ids,
            start_date=filters.start_date,
            end_date=filters.end_date,
        )

        return report_data

    def generate_csv_report(
        self,
        report_data: dict,
        filters: ReportFilter,
    ) -> BytesIO:
        """
        Genera un reporte en formato CSV.

        Args:
            report_data: Datos del reporte
            filters: Filtros utilizados

        Returns:
            BytesIO con contenido CSV
        """
        import csv

        output = BytesIO()
        
        # Usar encoding utf-8 con BOM para Excel
        text_output = []
        
        # Encabezado del reporte
        text_output.append("REPORTE DEPORTIVO")
        text_output.append(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        text_output.append("")

        # Resumen
        if report_data.get("athletes"):
            text_output.append(f"Total de Atletas: {len(report_data['athletes'])}")
        text_output.append(f"Asistencias: {report_data['statistics'].get('total_attendance', 0)}")
        text_output.append(f"Evaluaciones: {report_data['statistics'].get('total_evaluations', 0)}")
        text_output.append(f"Pruebas: {report_data['statistics'].get('total_tests', 0)}")
        text_output.append("")

        # Datos de asistencia
        if report_data.get("attendance"):
            text_output.append("=== ASISTENCIA ===")
            text_output.append("Atleta,Fecha,Hora,Estado")
            for record in report_data["attendance"]:
                athlete = next(
                    (a for a in report_data["athletes"] if a.id == record.athlete_id),
                    None,
                )
                if athlete:
                    text_output.append(
                        f"{athlete.full_name},{record.attendance_date},"
                        f"{record.attendance_time or ''},PRESENTE"
                    )
            text_output.append("")

        # Datos de evaluaciones
        if report_data.get("evaluations"):
            text_output.append("=== EVALUACIONES ===")
            text_output.append("Atleta,Fecha,Tipo,Puntuación,Comentarios")
            for record in report_data["evaluations"]:
                athlete = next(
                    (a for a in report_data["athletes"] if a.id == record.athlete_id),
                    None,
                )
                if athlete:
                    text_output.append(
                        f"{athlete.full_name},{record.created_at},"
                        f"{record.evaluation_type or ''},"
                        f"{record.score or ''},"
                        f"{record.comments or ''}"
                    )
            text_output.append("")

        # Datos de pruebas
        if report_data["tests"].get("sprint"):
            text_output.append("=== PRUEBAS DE VELOCIDAD ===")
            text_output.append("Atleta,Fecha,Distancia (m),Tiempo (s)")
            for record in report_data["tests"]["sprint"]:
                athlete = next(
                    (a for a in report_data["athletes"] if a.id == record.athlete_id),
                    None,
                )
                if athlete:
                    text_output.append(
                        f"{athlete.full_name},{record.created_at},"
                        f"{record.distance or ''},"
                        f"{record.time or ''}"
                    )

        # Escribir en BytesIO
        content = "\n".join(text_output)
        output.write(content.encode("utf-8-sig"))
        output.seek(0)

        return output

    def generate_xlsx_report(
        self,
        report_data: dict,
        filters: ReportFilter,
    ) -> BytesIO:
        """
        Genera un reporte en formato XLSX.

        Args:
            report_data: Datos del reporte
            filters: Filtros utilizados

        Returns:
            BytesIO con contenido XLSX
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ValidationException(
                "Se requiere openpyxl para generar reportes XLSX. "
                "Instale: pip install openpyxl"
            )

        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)

        # Hoja de Resumen
        ws_summary = workbook.create_sheet("Resumen", 0)
        ws_summary["A1"] = "REPORTE DEPORTIVO"
        ws_summary["A1"].font = Font(bold=True, size=14)
        ws_summary["A2"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws_summary["A4"] = "Estadísticas:"
        ws_summary["A5"] = f"Total de Atletas: {len(report_data.get('athletes', []))}"
        ws_summary["A6"] = f"Asistencias: {report_data['statistics'].get('total_attendance', 0)}"
        ws_summary["A7"] = f"Evaluaciones: {report_data['statistics'].get('total_evaluations', 0)}"
        ws_summary["A8"] = f"Pruebas: {report_data['statistics'].get('total_tests', 0)}"

        # Hoja de Asistencia
        if report_data.get("attendance"):
            ws_attendance = workbook.create_sheet("Asistencia")
            headers = ["Atleta", "Fecha", "Hora", "Estado"]
            ws_attendance.append(headers)

            for record in report_data["attendance"]:
                athlete = next(
                    (a for a in report_data["athletes"] if a.id == record.athlete_id),
                    None,
                )
                if athlete:
                    ws_attendance.append(
                        [
                            athlete.full_name,
                            str(record.attendance_date),
                            record.attendance_time or "",
                            "PRESENTE",
                        ]
                    )

        # Hoja de Evaluaciones
        if report_data.get("evaluations"):
            ws_eval = workbook.create_sheet("Evaluaciones")
            headers = ["Atleta", "Fecha", "Tipo", "Puntuación", "Comentarios"]
            ws_eval.append(headers)

            for record in report_data["evaluations"]:
                athlete = next(
                    (a for a in report_data["athletes"] if a.id == record.athlete_id),
                    None,
                )
                if athlete:
                    ws_eval.append(
                        [
                            athlete.full_name,
                            str(record.created_at),
                            record.evaluation_type or "",
                            record.score or "",
                            record.comments or "",
                        ]
                    )

        # Guardar en BytesIO
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        return output

    def generate_pdf_report(
        self,
        report_data: dict,
        filters: ReportFilter,
    ) -> BytesIO:
        """
        Genera un reporte en formato PDF.

        Args:
            report_data: Datos del reporte
            filters: Filtros utilizados

        Returns:
            BytesIO con contenido PDF
        """
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
        except ImportError:
            raise ValidationException(
                "Se requiere reportlab para generar reportes PDF. "
                "Instale: pip install reportlab"
            )

        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        # Título
        title = Paragraph(
            "REPORTE DEPORTIVO",
            ParagraphStyle(
                "CustomTitle",
                parent=styles["Heading1"],
                fontSize=18,
                textColor=colors.HexColor("#1f4788"),
                spaceAfter=12,
                alignment=1,
            ),
        )
        story.append(title)

        # Fecha
        date_text = Paragraph(
            f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
            styles["Normal"],
        )
        story.append(date_text)
        story.append(Spacer(1, 12))

        # Estadísticas
        stats_data = [
            ["ESTADÍSTICAS", ""],
            ["Total de Atletas", str(len(report_data.get("athletes", [])))],
            ["Asistencias", str(report_data["statistics"].get("total_attendance", 0))],
            ["Evaluaciones", str(report_data["statistics"].get("total_evaluations", 0))],
            ["Pruebas", str(report_data["statistics"].get("total_tests", 0))],
        ]

        stats_table = Table(stats_data, colWidths=[2 * inch, 2 * inch])
        stats_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4788")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        story.append(stats_table)
        story.append(Spacer(1, 20))

        # Tabla de Asistencia
        if report_data.get("attendance"):
            attendance_title = Paragraph(
                "Asistencia",
                ParagraphStyle(
                    "SubHeading",
                    parent=styles["Heading2"],
                    fontSize=12,
                    textColor=colors.HexColor("#1f4788"),
                ),
            )
            story.append(attendance_title)

            attendance_data = [["Atleta", "Fecha", "Hora", "Estado"]]
            for record in report_data["attendance"][:10]:  # Limitar a 10 para PDF
                athlete = next(
                    (a for a in report_data["athletes"] if a.id == record.athlete_id),
                    None,
                )
                if athlete:
                    attendance_data.append(
                        [
                            athlete.full_name[:20],
                            str(record.attendance_date),
                            record.attendance_time or "-",
                            "PRESENTE",
                        ]
                    )

            attendance_table = Table(
                attendance_data, colWidths=[2 * inch, 1.5 * inch, 1.5 * inch, 1 * inch]
            )
            attendance_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4788")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                        ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                    ]
                )
            )
            story.append(attendance_table)
            story.append(Spacer(1, 12))

        # Generar PDF
        doc.build(story)
        output.seek(0)

        return output

    def generate_report(
        self,
        db: Session,
        filters: ReportFilter,
        user_dni: str,
    ) -> ReportGenerationResponse:
        """
        Genera un reporte completo con los datos especificados.

        Args:
            db: Sesión de BD
            filters: Filtros del reporte
            user_dni: DNI del usuario que genera el reporte

        Returns:
            Respuesta con información del reporte generado
        """
        try:
            # Recolectar datos
            report_data = self.collect_report_data(db=db, filters=filters)

            # Generar archivo según formato
            report_id = str(uuid.uuid4())
            filename = f"reporte_deportivo_{report_id[:8]}"

            if filters.format == "csv":
                file_content = self.generate_csv_report(report_data, filters)
                filename += ".csv"
                mime_type = "text/csv"
            elif filters.format == "xlsx":
                file_content = self.generate_xlsx_report(report_data, filters)
                filename += ".xlsx"
                mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            elif filters.format == "pdf":
                file_content = self.generate_pdf_report(report_data, filters)
                filename += ".pdf"
                mime_type = "application/pdf"
            else:
                raise ValidationException(f"Formato no soportado: {filters.format}")

            file_size = len(file_content.getvalue())

            logger.info(
                f"Reporte generado exitosamente. ID: {report_id}, "
                f"Formato: {filters.format}, Tamaño: {file_size} bytes"
            )

            return ReportGenerationResponse(
                report_id=report_id,
                format=filters.format,
                file_name=filename,
                file_size=file_size,
                generated_at=datetime.now().isoformat(),
                expires_at=(
                    (datetime.now().timestamp() + 86400) if file_size < 104857600 else None
                ),
            )

        except Exception as e:
            logger.error(f"Error generando reporte: {str(e)}")
            raise ValidationException(f"Error al generar reporte: {str(e)}")
