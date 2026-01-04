"""Servicio de reportes que reutiliza lógica de estadísticas."""

import csv
import logging
from io import BytesIO, StringIO

import openpyxl
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

from app.controllers.statistic_controller import StatisticController
from app.schemas.report_schema import ReportFilter, ReportMetadata, ReportType

logger = logging.getLogger(__name__)


class ReportService:
    """Servicio para generación de reportes reutilizando lógica de estadísticas."""

    def __init__(self):
        self.statistic_controller = StatisticController()

    def generate_attendance_report(
        self,
        db: Session,
        filters: ReportFilter,
        user_name: str,
    ) -> BytesIO:
        """Genera reporte de asistencia."""
        stats_data = self.statistic_controller.get_attendance_statistics(
            db=db,
            start_date=filters.start_date,
            end_date=filters.end_date,
            type_athlete=filters.athlete_type.value if filters.athlete_type else None,
            sex=filters.sex.value if filters.sex else None,
        )

        metadata = self._build_metadata(
            title="Reporte de Asistencia",
            filters=filters,
            user_name=user_name,
            report_type=ReportType.ATTENDANCE,
        )

        if filters.format == "pdf":
            return self._generate_attendance_pdf(stats_data, metadata)
        elif filters.format == "xlsx":
            return self._generate_attendance_excel(stats_data, metadata)
        else:  # csv
            return self._generate_attendance_csv(stats_data, metadata)

    def generate_tests_report(
        self,
        db: Session,
        filters: ReportFilter,
        user_name: str,
    ) -> BytesIO:
        """Genera reporte de evaluaciones y tests."""
        stats_data = self.statistic_controller.get_test_performance(
            db=db,
            start_date=filters.start_date,
            end_date=filters.end_date,
            type_athlete=filters.athlete_type.value if filters.athlete_type else None,
        )

        metadata = self._build_metadata(
            title="Reporte de Evaluaciones y Tests",
            filters=filters,
            user_name=user_name,
            report_type=ReportType.TESTS,
        )

        if filters.format == "pdf":
            return self._generate_tests_pdf(stats_data, metadata)
        elif filters.format == "xlsx":
            return self._generate_tests_excel(stats_data, metadata)
        else:
            return self._generate_tests_csv(stats_data, metadata)

    def generate_statistics_report(
        self,
        db: Session,
        filters: ReportFilter,
        user_name: str,
    ) -> BytesIO:
        """Genera reporte de estadísticas generales."""
        stats_data = self.statistic_controller.get_club_overview(
            db=db,
            type_athlete=filters.athlete_type.value if filters.athlete_type else None,
            sex=filters.sex.value if filters.sex else None,
        )

        metadata = self._build_metadata(
            title="Reporte de Estadísticas Generales",
            filters=filters,
            user_name=user_name,
            report_type=ReportType.STATISTICS,
        )

        if filters.format == "pdf":
            return self._generate_statistics_pdf(stats_data, metadata)
        elif filters.format == "xlsx":
            return self._generate_statistics_excel(stats_data, metadata)
        else:
            return self._generate_statistics_csv(stats_data, metadata)

    def _build_metadata(
        self,
        title: str,
        filters: ReportFilter,
        user_name: str,
        report_type: ReportType,
    ) -> ReportMetadata:
        """Construye metadata formal para el reporte."""

        period = None
        if filters.start_date and filters.end_date:
            if filters.start_date.month == filters.end_date.month:
                period = filters.start_date.strftime("%B %Y")
            else:
                start_str = filters.start_date.strftime("%d/%m/%Y")
                end_str = filters.end_date.strftime("%d/%m/%Y")
                period = f"{start_str} - {end_str}"

        filters_dict = {}

        if filters.athlete_id:
            try:
                filters_dict["Deportista ID"] = str(filters.athlete_id)
            except Exception:
                filters_dict["Deportista ID"] = str(filters.athlete_id)

        if filters.athlete_type:
            filters_dict["Tipo de Deportista"] = filters.athlete_type.value
        if filters.sex:
            filters_dict["Sexo"] = filters.sex.value
        if filters.start_date:
            filters_dict["Fecha Inicio"] = filters.start_date.strftime("%d/%m/%Y")
        if filters.end_date:
            filters_dict["Fecha Fin"] = filters.end_date.strftime("%d/%m/%Y")

        return ReportMetadata(
            title=title,
            generated_by=user_name,
            filters_applied=filters_dict,
            period=period,
        )

    # ==================== PDF ====================

    def _generate_attendance_pdf(
        self, stats_data: dict, metadata: ReportMetadata
    ) -> BytesIO:
        """Genera reporte de asistencia en PDF."""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        elements.extend(self._build_pdf_header(metadata, styles))

        # Tabla por tipo de deportista
        if "attendance_by_type" in stats_data and stats_data["attendance_by_type"]:
            elements.append(Spacer(1, 0.3 * inch))
            elements.append(
                Paragraph("Asistencia por Tipo de Deportista", styles["Heading2"])
            )

            table_data = [["Tipo", "Total Registros", "Presentes", "% Asistencia"]]
            for item in stats_data["attendance_by_type"]:
                table_data.append(
                    [
                        item["type_athlete"],
                        str(item["total"]),
                        str(item["present"]),
                        f"{item.get('attendance_rate', 0):.1f}%",
                    ]
                )

            table = Table(table_data)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                    ]
                )
            )
            elements.append(table)

        # Resumen
        elements.append(Spacer(1, 0.3 * inch))
        elements.append(Paragraph("Resumen Global", styles["Heading2"]))
        resumen_data = [
            ["Total Registros", str(stats_data.get("total_records", 0))],
            ["Total Presentes", str(stats_data.get("total_present", 0))],
            ["Total Ausentes", str(stats_data.get("total_absent", 0))],
            [
                "Promedio Asistencia",
                f"{stats_data.get('overall_attendance_rate', 0):.1f}%",
            ],
        ]
        resumen_table = Table(resumen_data)
        resumen_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), colors.lightgrey),
                    ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                ]
            )
        )
        elements.append(resumen_table)

        doc.build(elements)
        buffer.seek(0)
        return buffer

    def _generate_tests_pdf(
        self, stats_data: dict, metadata: ReportMetadata
    ) -> BytesIO:
        """Genera reporte de tests en PDF."""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        elements.extend(self._build_pdf_header(metadata, styles))

        elements.append(Spacer(1, 0.3 * inch))
        elements.append(Paragraph("Rendimiento en Tests", styles["Heading2"]))

        if "tests_by_type" in stats_data and stats_data["tests_by_type"]:
            table_data = [["Tipo de Test", "Total Tests", "Promedio", "Mín", "Máx"]]
            for item in stats_data["tests_by_type"]:
                avg = (
                    f"{item.get('avg_score', 0):.2f}"
                    if item.get("avg_score")
                    else "N/A"
                )
                min_val = (
                    f"{item.get('min_score', 0):.2f}"
                    if item.get("min_score")
                    else "N/A"
                )
                max_val = (
                    f"{item.get('max_score', 0):.2f}"
                    if item.get("max_score")
                    else "N/A"
                )
                table_data.append(
                    [
                        item["test_type"],
                        str(item["total_tests"]),
                        avg,
                        min_val,
                        max_val,
                    ]
                )

            table = Table(table_data)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                    ]
                )
            )
            elements.append(table)

        # Top performers
        if "top_performers" in stats_data and stats_data["top_performers"]:
            elements.append(Spacer(1, 0.3 * inch))
            elements.append(
                Paragraph("Top 5 - Deportistas más evaluados", styles["Heading2"])
            )

            performers_data = [["Deportista", "Tipo", "Tests Completados"]]
            for performer in stats_data["top_performers"]:
                performers_data.append(
                    [
                        performer["athlete_name"],
                        performer["athlete_type"],
                        str(performer["tests_completed"]),
                    ]
                )

            performers_table = Table(performers_data)
            performers_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16a34a")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                    ]
                )
            )
            elements.append(performers_table)

        doc.build(elements)
        buffer.seek(0)
        return buffer

    def _generate_statistics_pdf(
        self, stats_data: dict, metadata: ReportMetadata
    ) -> BytesIO:
        """Genera reporte de estadísticas generales en PDF."""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        elements.extend(self._build_pdf_header(metadata, styles))

        # Estadísticas generales
        elements.append(Spacer(1, 0.3 * inch))
        elements.append(Paragraph("Resumen General del Club", styles["Heading2"]))

        table_data = [
            ["Métrica", "Valor"],
            ["Total Deportistas", str(stats_data.get("total_athletes", 0))],
            ["Deportistas Activos", str(stats_data.get("active_athletes", 0))],
            ["Deportistas Inactivos", str(stats_data.get("inactive_athletes", 0))],
            ["Total Evaluaciones", str(stats_data.get("total_evaluations", 0))],
            ["Total Tests Completados", str(stats_data.get("total_tests", 0))],
        ]

        table = Table(table_data)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                ]
            )
        )
        elements.append(table)

        # Distribución por tipo
        if "athletes_by_type" in stats_data and stats_data["athletes_by_type"]:
            elements.append(Spacer(1, 0.3 * inch))
            elements.append(Paragraph("Distribución por Tipo", styles["Heading2"]))

            type_data = [["Tipo de Deportista", "Cantidad", "Porcentaje"]]
            for item in stats_data["athletes_by_type"]:
                type_data.append(
                    [
                        item["type_athlete"],
                        str(item["count"]),
                        f"{item.get('percentage', 0):.1f}%",
                    ]
                )

            type_table = Table(type_data)
            type_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#7c3aed")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                    ]
                )
            )
            elements.append(type_table)

        # Distribución por género
        if "athletes_by_gender" in stats_data and stats_data["athletes_by_gender"]:
            elements.append(Spacer(1, 0.3 * inch))
            elements.append(Paragraph("Distribución por Género", styles["Heading2"]))

            gender_data = [["Género", "Cantidad", "Porcentaje"]]
            for item in stats_data["athletes_by_gender"]:
                gender_data.append(
                    [
                        item["sex"],
                        str(item["count"]),
                        f"{item.get('percentage', 0):.1f}%",
                    ]
                )

            gender_table = Table(gender_data)
            gender_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dc2626")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                    ]
                )
            )
            elements.append(gender_table)

        doc.build(elements)
        buffer.seek(0)
        return buffer

    def _build_pdf_header(self, metadata: ReportMetadata, styles) -> list:
        """Construye header profesional mejorado para PDFs."""
        from reportlab.platypus import HRFlowable

        elements = []

        # Línea superior decorativa
        elements.append(
            HRFlowable(width="100%", thickness=2, color=colors.HexColor("#1e40af"))
        )
        elements.append(Spacer(1, 0.1 * inch))

        # Sistema - Centrado y destacado
        system_style = ParagraphStyle(
            "SystemName",
            parent=styles["Normal"],
            fontSize=16,
            textColor=colors.HexColor("#1e40af"),
            alignment=1,  # Center
            fontName="Helvetica-Bold",
            spaceAfter=4,
        )
        elements.append(Paragraph("SISTEMA KALLPA UNL", system_style))

        # Título del reporte - Centrado
        title_style = ParagraphStyle(
            "ReportTitle",
            parent=styles["Heading1"],
            fontSize=14,
            textColor=colors.HexColor("#3b82f6"),
            alignment=1,  # Center
            spaceAfter=12,
        )
        elements.append(Paragraph(metadata.title, title_style))

        # Línea separadora
        elements.append(HRFlowable(width="100%", thickness=1, color=colors.lightgrey))
        elements.append(Spacer(1, 0.15 * inch))

        # Metadata en formato tabla para mejor alineación
        meta_data = []
        meta_data.append(
            ["Generado:", metadata.generated_at.strftime("%d/%m/%Y %H:%M")]
        )
        meta_data.append(["Por:", metadata.generated_by])

        if metadata.period:
            meta_data.append(["Período:", metadata.period])

        meta_table = Table(meta_data, colWidths=[1.2 * inch, 4 * inch])
        meta_table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#64748b")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ]
            )
        )
        elements.append(meta_table)

        # Filtros aplicados si existen
        if metadata.filters_applied:
            elements.append(Spacer(1, 0.15 * inch))

            filter_title_style = ParagraphStyle(
                "FilterTitle",
                parent=styles["Normal"],
                fontSize=10,
                fontName="Helvetica-Bold",
                textColor=colors.HexColor("#64748b"),
            )
            elements.append(Paragraph("Filtros Aplicados:", filter_title_style))
            elements.append(Spacer(1, 0.05 * inch))

            filter_data = [
                [key, str(value)] for key, value in metadata.filters_applied.items()
            ]
            filter_table = Table(filter_data, colWidths=[1.8 * inch, 3.5 * inch])
            filter_table.setStyle(
                TableStyle(
                    [
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#64748b")),
                        ("LEFTPADDING", (0, 0), (-1, -1), 0),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                        ("TOPPADDING", (0, 0), (-1, -1), 2),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                    ]
                )
            )
            elements.append(filter_table)

        # Espaciado final antes del contenido
        elements.append(Spacer(1, 0.3 * inch))

        # Línea decorativa antes del contenido
        elements.append(
            HRFlowable(width="100%", thickness=1, color=colors.HexColor("#e5e7eb"))
        )
        elements.append(Spacer(1, 0.2 * inch))

        return elements

    # ==================== EXCEL ====================

    def _generate_attendance_excel(
        self, stats_data: dict, metadata: ReportMetadata
    ) -> BytesIO:
        """Genera reporte de asistencia en Excel."""
        workbook = openpyxl.Workbook()
        metadata_sheet = workbook.active
        metadata_sheet.title = "Información"
        self._write_excel_metadata(metadata_sheet, metadata)

        data_sheet = workbook.create_sheet("Asistencias")
        self._write_attendance_data_excel(data_sheet, stats_data)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    def _generate_tests_excel(
        self, stats_data: dict, metadata: ReportMetadata
    ) -> BytesIO:
        """Genera reporte de tests en Excel."""
        workbook = openpyxl.Workbook()
        metadata_sheet = workbook.active
        metadata_sheet.title = "Información"
        self._write_excel_metadata(metadata_sheet, metadata)

        data_sheet = workbook.create_sheet("Tests")
        self._write_tests_data_excel(data_sheet, stats_data)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    def _generate_statistics_excel(
        self, stats_data: dict, metadata: ReportMetadata
    ) -> BytesIO:
        """Genera reporte de estadísticas en Excel."""
        workbook = openpyxl.Workbook()
        metadata_sheet = workbook.active
        metadata_sheet.title = "Información"
        self._write_excel_metadata(metadata_sheet, metadata)

        data_sheet = workbook.create_sheet("Estadísticas")
        self._write_statistics_data_excel(data_sheet, stats_data)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    def _write_excel_metadata(self, sheet, metadata: ReportMetadata):
        """Escribe metadata en hoja de Excel."""
        sheet["A1"] = metadata.title
        sheet["A1"].font = Font(size=14, bold=True)

        sheet["A3"] = "Sistema:"
        sheet["B3"] = metadata.system_name

        sheet["A4"] = "Generado:"
        sheet["B4"] = metadata.generated_at.strftime("%d/%m/%Y %H:%M")

        sheet["A5"] = "Por:"
        sheet["B5"] = metadata.generated_by

        if metadata.period:
            sheet["A6"] = "Período:"
            sheet["B6"] = metadata.period

        row = 8
        if metadata.filters_applied:
            sheet[f"A{row}"] = "Filtros Aplicados:"
            sheet[f"A{row}"].font = Font(bold=True)
            row += 1
            for key, value in metadata.filters_applied.items():
                sheet[f"A{row}"] = key
                sheet[f"B{row}"] = value
                row += 1

    def _write_attendance_data_excel(self, sheet, stats_data: dict):
        """Escribe datos de asistencia en Excel."""
        headers = ["Tipo", "Total Registros", "Presentes", "% Asistencia"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="1e40af", end_color="1e40af", fill_type="solid"
            )

        if "attendance_by_type" in stats_data:
            for row, item in enumerate(stats_data["attendance_by_type"], 2):
                sheet.cell(row=row, column=1, value=item["type_athlete"])
                sheet.cell(row=row, column=2, value=item["total"])
                sheet.cell(row=row, column=3, value=item["present"])
                sheet.cell(
                    row=row, column=4, value=f"{item.get('attendance_rate', 0):.1f}%"
                )

    def _write_tests_data_excel(self, sheet, stats_data: dict):
        """Escribe datos de tests en Excel."""
        headers = ["Tipo de Test", "Total", "Promedio", "Mín", "Máx"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        if "tests_by_type" in stats_data:
            for row, item in enumerate(stats_data["tests_by_type"], 2):
                sheet.cell(row=row, column=1, value=item["test_type"])
                sheet.cell(row=row, column=2, value=item["total_tests"])
                sheet.cell(
                    row=row,
                    column=3,
                    value=round(item["avg_score"], 2)
                    if item.get("avg_score")
                    else "N/A",
                )
                sheet.cell(
                    row=row,
                    column=4,
                    value=round(item["min_score"], 2)
                    if item.get("min_score")
                    else "N/A",
                )
                sheet.cell(
                    row=row,
                    column=5,
                    value=round(item["max_score"], 2)
                    if item.get("max_score")
                    else "N/A",
                )

    def _write_statistics_data_excel(self, sheet, stats_data: dict):
        """Escribe estadísticas generales en Excel."""
        headers = ["Métrica", "Valor"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        metrics = [
            ("Total Deportistas", stats_data.get("total_athletes", 0)),
            ("Activos", stats_data.get("active_athletes", 0)),
            ("Inactivos", stats_data.get("inactive_athletes", 0)),
            ("Total Evaluaciones", stats_data.get("total_evaluations", 0)),
            ("Total Tests", stats_data.get("total_tests", 0)),
        ]

        for row, (metric, value) in enumerate(metrics, 2):
            sheet.cell(row=row, column=1, value=metric)
            sheet.cell(row=row, column=2, value=value)

    # ==================== CSV ====================

    def _generate_attendance_csv(
        self, stats_data: dict, metadata: ReportMetadata
    ) -> BytesIO:
        """Genera reporte de asistencia en CSV."""
        output = StringIO()
        writer = csv.writer(output)

        writer.writerow([f"# {metadata.title}"])
        writer.writerow(
            [f"# Generado: {metadata.generated_at.strftime('%d/%m/%Y %H:%M')}"]
        )
        writer.writerow([])
        writer.writerow(["Tipo", "Total Registros", "Presentes", "% Asistencia"])

        if "attendance_by_type" in stats_data:
            for item in stats_data["attendance_by_type"]:
                writer.writerow(
                    [
                        item["type_athlete"],
                        item["total"],
                        item["present"],
                        f"{item.get('attendance_rate', 0):.1f}%",
                    ]
                )

        buffer = BytesIO()
        buffer.write(output.getvalue().encode("utf-8-sig"))
        buffer.seek(0)
        return buffer

    def _generate_tests_csv(
        self, stats_data: dict, metadata: ReportMetadata
    ) -> BytesIO:
        """Genera reporte de tests en CSV."""
        output = StringIO()
        writer = csv.writer(output)

        writer.writerow([f"# {metadata.title}"])
        writer.writerow([])
        writer.writerow(["Tipo Test", "Total", "Promedio", "Mín", "Máx"])

        if "tests_by_type" in stats_data:
            for item in stats_data["tests_by_type"]:
                avg = (
                    f"{item.get('avg_score', 0):.2f}"
                    if item.get("avg_score")
                    else "N/A"
                )
                min_val = (
                    f"{item.get('min_score', 0):.2f}"
                    if item.get("min_score")
                    else "N/A"
                )
                max_val = (
                    f"{item.get('max_score', 0):.2f}"
                    if item.get("max_score")
                    else "N/A"
                )
                writer.writerow(
                    [
                        item["test_type"],
                        item["total_tests"],
                        avg,
                        min_val,
                        max_val,
                    ]
                )

        buffer = BytesIO()
        buffer.write(output.getvalue().encode("utf-8-sig"))
        buffer.seek(0)
        return buffer

    def _generate_statistics_csv(
        self, stats_data: dict, metadata: ReportMetadata
    ) -> BytesIO:
        """Genera reporte de estadísticas en CSV."""
        output = StringIO()
        writer = csv.writer(output)

        writer.writerow([f"# {metadata.title}"])
        writer.writerow([])
        writer.writerow(["Métrica", "Valor"])

        writer.writerow(["Total Deportistas", stats_data.get("total_athletes", 0)])
        writer.writerow(["Activos", stats_data.get("active_athletes", 0)])
        writer.writerow(["Inactivos", stats_data.get("inactive_athletes", 0)])
        writer.writerow(["Total Evaluaciones", stats_data.get("total_evaluations", 0)])
        writer.writerow(["Total Tests", stats_data.get("total_tests", 0)])

        buffer = BytesIO()
        buffer.write(output.getvalue().encode("utf-8-sig"))
        buffer.seek(0)
        return buffer
